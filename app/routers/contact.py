from fastapi import APIRouter, status, HTTPException
from app import schemas
from typing import List
import os
from openpyxl import Workbook, load_workbook

router = APIRouter(prefix="/api/contact", tags=["contact"])

EXCEL_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'contacts_new.xlsx')
HEADERS = ['Name', 'Email', 'Message']

def ensure_excel_file():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)

def add_contact_to_excel(name: str, email: str, message: str):
    try:
        ensure_excel_file()
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([name, email, message])
        wb.save(EXCEL_FILE)
        print(f"Data appended to Excel: {name}, {email}")
    except PermissionError as e:
        print(f"Permission error in add_contact_to_excel: {str(e)}")
        print(f"Excel file path: {EXCEL_FILE}")
        print(f"File exists: {os.path.exists(EXCEL_FILE)}")
        raise e
    except Exception as e:
        print(f"Error in add_contact_to_excel: {str(e)}")
        print(f"Excel file path: {EXCEL_FILE}")
        print(f"File exists: {os.path.exists(EXCEL_FILE)}")
        raise e

def read_contacts_from_excel():
    ensure_excel_file()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    contacts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue
        contacts.append({
            'name': row[0],
            'email': row[1],
            'message': row[2],
        })
    return contacts

@router.post("/", response_model=schemas.ContactBase, status_code=status.HTTP_201_CREATED)
def submit_contact(form: schemas.ContactCreate):
    try:
        add_contact_to_excel(form.name, form.email, form.message)
        print(f"Contact saved successfully: {form.name}, {form.email}")
        return form
    except PermissionError as e:
        print(f"Permission error saving contact: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Unable to save contact. Please ensure the Excel file is not open in another program."
        )
    except Exception as e:
        print(f"Error saving contact: {str(e)}")
        print(f"Excel file path: {EXCEL_FILE}")
        print(f"File exists: {os.path.exists(EXCEL_FILE)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error saving contact: {str(e)}"
        )

@router.get("/", response_model=List[schemas.ContactBase])
def list_contacts():
    return read_contacts_from_excel() 